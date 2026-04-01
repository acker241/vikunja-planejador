#!/usr/bin/env python3
"""
Importador de tarefas do Manual Operacional para Vikunja.

Lê o Excel 'Manual_Operacional_Associacao_v1.xlsx' e cria projetos/tarefas
no Vikunja via API REST.

Uso:
    py scripts/import_vikunja.py --url http://localhost:3456 --user admin --password senha123

Requisitos:
    pip install openpyxl requests
"""

import argparse
import io
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path

# Fix Windows console encoding
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl
import requests


# ──────────────────────────────────────────────
# Configuracao
# ──────────────────────────────────────────────

EXCEL_FILE = Path(__file__).parent.parent / "Manual_Operacional_Associacao_v1.xlsx"

LABEL_COLORS = {
    "URGENTE": "#e74c3c",
    "ALTA": "#e67e22",
    "MEDIA": "#f1c40f",
    "BAIXA": "#3498db",
    "BLOQUEADO": "#2c3e50",
    "CONCLUIDO": "#27ae60",
}

STATUS_MAP = {
    "Concluído": True,
    "🟢 Concluído": True,
    "🟢": True,
    "Pendente": False,
    "Em andamento": False,
    "🟡 Em andamento": False,
    "🟡": False,
    "Em Preparação": False,
    "🔴 Não iniciado": False,
    "🔴": False,
    "⚫ Bloqueado": False,
    "⚫": False,
    "🔵 Aguardando terceiro": False,
    "🔵": False,
}

PRIORITY_MAP = {
    "🔴 URGENTE": 5,
    "URGENTE": 5,
    "ALTA": 4,
    "MEDIA": 3,
    "MÉDIA": 3,
    "BAIXA": 2,
}


# ──────────────────────────────────────────────
# API Client
# ──────────────────────────────────────────────

class VikunjaClient:
    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.token = None

    def login(self, username: str, password: str):
        resp = self.session.post(f"{self.base_url}/api/v1/login", json={
            "username": username,
            "password": password,
        })
        resp.raise_for_status()
        self.token = resp.json()["token"]
        self.session.headers["Authorization"] = f"Bearer {self.token}"
        print(f"  Login OK: {username}")

    def register(self, username: str, email: str, password: str):
        resp = self.session.post(f"{self.base_url}/api/v1/register", json={
            "username": username,
            "email": email,
            "password": password,
        })
        if resp.status_code == 200:
            print(f"  Utilizador criado: {username}")
            return resp.json()
        elif resp.status_code == 400 and "already exists" in resp.text.lower():
            print(f"  Utilizador ja existe: {username}")
            return None
        else:
            print(f"  Aviso ao criar {username}: {resp.status_code} {resp.text[:200]}")
            return None

    def create_project(self, title: str, description: str = "", parent_id: int = 0) -> dict:
        payload = {"title": title, "description": description}
        if parent_id:
            payload["parent_project_id"] = parent_id
        resp = self.session.put(f"{self.base_url}/api/v1/projects", json=payload)
        resp.raise_for_status()
        proj = resp.json()
        print(f"  Projeto criado: {title} (id={proj['id']})")
        return proj

    def create_task(self, project_id: int, title: str, description: str = "",
                    done: bool = False, priority: int = 0,
                    due_date: str = None, labels: list = None,
                    bucket_id: int = None) -> dict:
        payload = {
            "title": title[:250],
            "description": description,
            "done": done,
            "priority": priority,
        }
        if due_date:
            payload["due_date"] = due_date
        if labels:
            payload["labels"] = labels
        if bucket_id:
            payload["bucket_id"] = bucket_id

        resp = self.session.put(
            f"{self.base_url}/api/v1/projects/{project_id}/tasks",
            json=payload,
        )
        resp.raise_for_status()
        return resp.json()

    def create_label(self, title: str, hex_color: str = "#cecece") -> dict:
        resp = self.session.put(f"{self.base_url}/api/v1/labels", json={
            "title": title,
            "hex_color": hex_color,
        })
        resp.raise_for_status()
        label = resp.json()
        print(f"  Label criada: {title} ({hex_color})")
        return label

    def get_labels(self) -> list:
        resp = self.session.get(f"{self.base_url}/api/v1/labels")
        resp.raise_for_status()
        return resp.json() or []

    def create_bucket(self, project_id: int, title: str) -> dict:
        resp = self.session.put(
            f"{self.base_url}/api/v1/projects/{project_id}/buckets",
            json={"title": title},
        )
        resp.raise_for_status()
        return resp.json()

    def get_buckets(self, project_id: int) -> list:
        resp = self.session.get(f"{self.base_url}/api/v1/projects/{project_id}/buckets")
        resp.raise_for_status()
        return resp.json() or []


# ──────────────────────────────────────────────
# Excel Parsers
# ──────────────────────────────────────────────

def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def parse_tarefas_por_pessoa(wb) -> list:
    """Extrai tarefas da aba 16_TAREFAS_POR_PESSOA."""
    ws = wb["16_TAREFAS_POR_PESSOA"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        # Detectar linhas de tarefa: começam com ID tipo C-01 ou J-01
        row_list = list(row)
        if len(row_list) >= 11:
            cell_id = safe_str(row_list[1])
            if cell_id and ("-" in cell_id) and len(cell_id) <= 5:
                task_id = cell_id
                responsavel = safe_str(row_list[2])
                tarefa = safe_str(row_list[3])
                prazo = safe_str(row_list[4])
                status = safe_str(row_list[5])
                prioridade = safe_str(row_list[6])
                entregavel = safe_str(row_list[7])
                depende_de = safe_str(row_list[8])
                obs = safe_str(row_list[9])

                # Extrair titulo (primeira linha da tarefa)
                titulo_lines = tarefa.split("\n")
                titulo = f"[{task_id}] {titulo_lines[0]}"
                detalhes = "\n".join(titulo_lines[1:]) if len(titulo_lines) > 1 else ""

                descricao_parts = []
                if responsavel:
                    descricao_parts.append(f"**Responsavel:** {responsavel}")
                if prazo:
                    descricao_parts.append(f"**Prazo:** {prazo}")
                if detalhes:
                    descricao_parts.append(f"\n**Detalhes:**\n{detalhes}")
                if entregavel:
                    descricao_parts.append(f"\n**Entregavel:**\n{entregavel}")
                if depende_de and depende_de != "—":
                    descricao_parts.append(f"\n**Depende de:** {depende_de}")
                if obs:
                    descricao_parts.append(f"\n**Obs:** {obs}")

                done = STATUS_MAP.get(status, False)
                prio = PRIORITY_MAP.get(prioridade, 0)

                # Inferir labels
                labels = []
                if responsavel:
                    labels.append(responsavel)
                if prioridade in ("🔴 URGENTE", "URGENTE"):
                    labels.append("URGENTE")
                elif "ALTA" in prioridade.upper():
                    labels.append("ALTA")
                elif "MEDIA" in prioridade.upper() or "MÉDIA" in prioridade.upper():
                    labels.append("MEDIA")
                if "Bloqueado" in status or "⚫" in status:
                    labels.append("BLOQUEADO")

                tasks.append({
                    "title": titulo,
                    "description": "\n".join(descricao_parts),
                    "done": done,
                    "priority": prio,
                    "labels": labels,
                    "status_text": status,
                    "responsavel": responsavel,
                })

    return tasks


def parse_cronograma(wb) -> list:
    """Extrai atividades da aba 02_CRONOGRAMA_GERAL."""
    ws = wb["02_CRONOGRAMA_GERAL"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) >= 10:
            num = safe_str(row_list[1])
            if num and num.isdigit():
                atividade = safe_str(row_list[2])
                responsavel = safe_str(row_list[3])
                prerequisito = safe_str(row_list[4])
                obs = safe_str(row_list[9])

                # Determinar semanas ativas
                semanas = []
                for i, sem in enumerate([5, 6, 7, 8], 1):
                    if row_list[sem]:
                        semanas.append(f"Sem {i}")

                descricao_parts = []
                if responsavel:
                    descricao_parts.append(f"**Responsavel:** {responsavel}")
                if prerequisito and prerequisito != "—":
                    descricao_parts.append(f"**Pre-requisito:** {prerequisito}")
                if semanas:
                    descricao_parts.append(f"**Periodo:** {', '.join(semanas)}")
                if obs:
                    descricao_parts.append(f"**Obs:** {obs}")

                tasks.append({
                    "title": f"{num}. {atividade}",
                    "description": "\n".join(descricao_parts),
                    "done": False,
                    "priority": 0,
                    "labels": [responsavel] if responsavel else [],
                })

    return tasks


def parse_checklist_cartorio(wb) -> list:
    """Extrai itens da aba 03_CHECKLIST_CARTORIO."""
    ws = wb["03_CHECKLIST_CARTORIO"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) >= 7:
            num = row_list[0]
            if num and isinstance(num, (int, float)):
                categoria = safe_str(row_list[1])
                item = safe_str(row_list[2])
                descricao = safe_str(row_list[3])
                status = safe_str(row_list[4])
                responsavel = safe_str(row_list[5])
                obs = safe_str(row_list[6])

                done = status.lower() in ("concluído", "concluido")

                descricao_parts = []
                if categoria:
                    descricao_parts.append(f"**Categoria:** {categoria}")
                if descricao:
                    descricao_parts.append(f"\n{descricao}")
                if responsavel and responsavel != "—":
                    descricao_parts.append(f"\n**Responsavel:** {responsavel}")
                if obs:
                    descricao_parts.append(f"\n**Obs:** {obs}")

                labels = []
                if responsavel and responsavel != "—":
                    labels.append(responsavel)
                if not done and status:
                    labels.append(status)

                tasks.append({
                    "title": f"#{int(num)} {item}",
                    "description": "\n".join(descricao_parts),
                    "done": done,
                    "priority": 0,
                    "labels": labels,
                })

    return tasks


def parse_pendencias(wb) -> list:
    """Extrai itens da aba 14_PENDENCIAS_INSUMO."""
    ws = wb["14_PENDENCIAS_INSUMO"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) >= 7:
            num_raw = safe_str(row_list[1])
            if not num_raw:
                continue
            try:
                num = int(float(num_raw))
            except (ValueError, TypeError):
                continue
            if num < 1 or num > 99:
                continue

            pendencia = safe_str(row_list[2])
            fonte = safe_str(row_list[3])
            criticidade = safe_str(row_list[4])
            responsavel = safe_str(row_list[5])
            obs = safe_str(row_list[6])
            if not pendencia:
                continue

            done = "CONCLU" in criticidade.upper() or "CONCLU" in obs.upper()[:20]

            descricao_parts = []
            if fonte:
                descricao_parts.append(f"**Fonte:** {fonte}")
            if criticidade:
                descricao_parts.append(f"**Criticidade:** {criticidade}")
            if responsavel:
                descricao_parts.append(f"**Responsavel:** {responsavel}")
            if obs:
                descricao_parts.append(f"\n**Obs:** {obs}")

            prio = 4 if "ALTA" in criticidade.upper() else 3
            labels = ["BLOQUEANTE"] if "BLOQ" in criticidade.upper() or "BLOQ" in obs.upper() else []
            if responsavel:
                labels.append(responsavel)

            tasks.append({
                "title": f"Pendencia #{num}: {pendencia}",
                    "description": "\n".join(descricao_parts),
                "done": done,
                "priority": prio,
                "labels": labels,
            })

    return tasks


def parse_riscos(wb) -> list:
    """Extrai itens da aba 12_RISCOS_CONTROLES."""
    ws = wb["12_RISCOS_CONTROLES"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) >= 8:
            num_raw = safe_str(row_list[1])
            if not num_raw:
                continue
            try:
                num = int(float(num_raw))
            except (ValueError, TypeError):
                continue
            if num < 1 or num > 99:
                continue

            risco = safe_str(row_list[2])
            if not risco or risco == "Risco":
                continue

            probabilidade = safe_str(row_list[3])
            impacto = safe_str(row_list[4])
            controle = safe_str(row_list[5])
            responsavel = safe_str(row_list[6])

            descricao_parts = [
                f"**Probabilidade:** {probabilidade}",
                f"**Impacto:** {impacto}",
                f"\n**Controle/Mitigacao:**\n{controle}",
            ]
            if responsavel:
                descricao_parts.append(f"\n**Responsavel:** {responsavel}")

            prio = 4 if impacto in ("Alto", "Critico") else 3
            labels = [f"Impacto {impacto}"]
            if responsavel:
                labels.append(responsavel)

            tasks.append({
                "title": f"Risco #{num}: {risco}",
                "description": "\n".join(descricao_parts),
                "done": False,
                "priority": prio,
                "labels": labels,
            })

    return tasks


def parse_cnpj_redesim(wb) -> list:
    """Extrai passos da aba 04_CNPJ_REDESIM_RFB."""
    ws = wb["04_CNPJ_REDESIM_RFB"]
    tasks = []
    current_phase = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) >= 7:
            passo = safe_str(row_list[1])

            # Detectar fases
            if passo.startswith("FASE"):
                current_phase = passo
                continue

            # Detectar passos numerados (1.1, 2.1, etc)
            if passo and "." in passo:
                try:
                    float(passo)
                except ValueError:
                    continue

                oque_fazer = safe_str(row_list[2])
                quem = safe_str(row_list[3])
                onde = safe_str(row_list[4])
                obs = safe_str(row_list[6])

                descricao_parts = []
                if current_phase:
                    descricao_parts.append(f"**Fase:** {current_phase}")
                if quem:
                    descricao_parts.append(f"**Quem:** {quem}")
                if onde:
                    descricao_parts.append(f"**Onde:** {onde}")
                if obs:
                    descricao_parts.append(f"**Obs:** {obs}")

                labels = []
                if quem:
                    labels.append(quem)

                tasks.append({
                    "title": f"[{passo}] {oque_fazer[:200]}",
                    "description": "\n".join(descricao_parts),
                    "done": False,
                    "priority": 0,
                    "labels": labels,
                })

    return tasks


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def ensure_labels(client: VikunjaClient, needed_labels: set) -> dict:
    """Cria labels que nao existem e retorna mapa nome->label_obj."""
    existing = {l["title"]: l for l in client.get_labels()}
    label_map = {}

    for name in needed_labels:
        if name in existing:
            label_map[name] = existing[name]
        else:
            color = LABEL_COLORS.get(name.upper(), "#95a5a6")
            label = client.create_label(name, color)
            label_map[name] = label

    return label_map


def import_tasks(client: VikunjaClient, project_id: int, tasks: list,
                 label_map: dict, project_name: str):
    """Importa lista de tarefas para um projeto."""
    count_done = 0
    count_todo = 0

    for t in tasks:
        labels_payload = []
        for lname in t.get("labels", []):
            if lname in label_map:
                labels_payload.append({"id": label_map[lname]["id"]})

        task = client.create_task(
            project_id=project_id,
            title=t["title"],
            description=t.get("description", ""),
            done=t.get("done", False),
            priority=t.get("priority", 0),
            labels=labels_payload if labels_payload else None,
        )
        if t.get("done"):
            count_done += 1
        else:
            count_todo += 1

    print(f"  {project_name}: {count_todo} pendentes, {count_done} concluidas ({len(tasks)} total)")


def main():
    parser = argparse.ArgumentParser(description="Importar tarefas do Excel para Vikunja")
    parser.add_argument("--url", required=True, help="URL do Vikunja (ex: http://localhost:3456)")
    parser.add_argument("--user", required=True, help="Username do admin")
    parser.add_argument("--password", required=True, help="Password do admin")
    parser.add_argument("--excel", default=str(EXCEL_FILE), help="Caminho do Excel")
    parser.add_argument("--dry-run", action="store_true", help="Apenas mostrar o que seria importado")
    args = parser.parse_args()

    # Carregar Excel
    print(f"\n1. Carregando Excel: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    print(f"   Abas encontradas: {len(wb.sheetnames)}")

    # Parsear todas as abas relevantes
    print("\n2. Extraindo tarefas...")

    sections = {
        "Tarefas Operacionais": parse_tarefas_por_pessoa(wb),
        "Cronograma Geral": parse_cronograma(wb),
        "Checklist Cartorio": parse_checklist_cartorio(wb),
        "Processo CNPJ-REDESIM": parse_cnpj_redesim(wb),
        "Pendencias e Bloqueios": parse_pendencias(wb),
        "Riscos e Controles": parse_riscos(wb),
    }

    total = sum(len(v) for v in sections.values())
    for name, tasks in sections.items():
        done = sum(1 for t in tasks if t.get("done"))
        print(f"   {name}: {len(tasks)} tarefas ({done} concluidas)")
    print(f"   TOTAL: {total} tarefas")

    # Coletar todas as labels necessarias
    all_labels = set()
    for tasks in sections.values():
        for t in tasks:
            all_labels.update(t.get("labels", []))
    print(f"\n   Labels necessarias: {len(all_labels)}")

    if args.dry_run:
        print("\n=== DRY RUN — nenhuma alteracao feita ===")
        print(f"\nResumo: {total} tarefas em {len(sections)} sub-projetos")
        print(f"Labels: {', '.join(sorted(all_labels))}")

        for name, tasks in sections.items():
            print(f"\n--- {name} ---")
            for t in tasks:
                status = "DONE" if t.get("done") else "TODO"
                prio = f"P{t.get('priority', 0)}" if t.get("priority") else ""
                print(f"  [{status}] {prio} {t['title'][:80]}")
        return

    # Conectar ao Vikunja
    print(f"\n3. Conectando ao Vikunja: {args.url}")
    client = VikunjaClient(args.url)
    client.login(args.user, args.password)

    # Criar labels
    print("\n4. Criando labels...")
    label_map = ensure_labels(client, all_labels)

    # Criar projeto principal
    print("\n5. Criando projetos...")
    main_project = client.create_project(
        title="Associacao Adquirentes — Mondrian BFabbriani",
        description="Manual Operacional para constituicao da Associacao de Adquirentes "
                    "do Empreendimento Edificio Residencial MONDRIAN (Itapema/SC). "
                    "Importado automaticamente do Excel v2.1.",
    )

    # Criar sub-projetos e importar tarefas
    print("\n6. Importando tarefas...")
    for name, tasks in sections.items():
        if not tasks:
            continue
        sub = client.create_project(
            title=name,
            description=f"Importado da planilha — {len(tasks)} tarefas",
            parent_id=main_project["id"],
        )
        import_tasks(client, sub["id"], tasks, label_map, name)

    print(f"\n{'='*50}")
    print(f"Importacao concluida!")
    print(f"  Projeto principal: {main_project['id']}")
    print(f"  Sub-projetos: {len(sections)}")
    print(f"  Tarefas importadas: {total}")
    print(f"\nAceda em: {args.url}")
    print(f"{'='*50}\n")


if __name__ == "__main__":
    main()
