#!/usr/bin/env python3
"""
Importador COMPLETO de tarefas do Manual Operacional para Leantime.

Le o Excel e cria estrutura rica no Leantime:
- Projetos com milestones
- Tarefas com dependencias (predecessor/successor)
- Descricoes HTML formatadas
- Prioridades, status, tags, datas para Gantt

Uso:
    py scripts/import_leantime.py --url https://DOMINIO --api-key KEY
    py scripts/import_leantime.py --url https://DOMINIO --api-key KEY --dry-run

Requisitos:
    pip install openpyxl requests
"""

import argparse
import io
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl
import requests

EXCEL_FILE = Path(__file__).parent.parent / "Manual_Operacional_Associacao_v1.xlsx"

# Leantime priority: 1=Critical, 2=High, 3=Medium, 4=Low, 5=Lowest
PRIORITY_MAP = {
    "🔴 URGENTE": 1,
    "URGENTE": 1,
    "ALTA": 2,
    "MEDIA": 3,
    "MÉDIA": 3,
    "BAIXA": 4,
}

# Leantime status: 3=New, 4=InProgress, 1=Blocked, 0=Done, -1=Archived
STATUS_MAP = {
    "🔴 Não iniciado": 3,
    "🔴": 3,
    "🟡 Em andamento": 4,
    "🟡": 4,
    "🟢 Concluído": 0,
    "🟢": 0,
    "⚫ Bloqueado": 1,
    "⚫": 1,
    "🔵 Aguardando terceiro": 1,
    "🔵": 1,
    "Concluído": 0,
    "Pendente": 3,
    "Em andamento": 4,
    "Em Preparação": 4,
}

# Data base para calcular datas (assembleia prevista 15/04/2026)
BASE_DATE = datetime(2026, 4, 1)
ASSEMBLEIA_DATE = datetime(2026, 4, 15)

WEEK_DATES = {
    1: (BASE_DATE, BASE_DATE + timedelta(days=6)),
    2: (BASE_DATE + timedelta(days=7), BASE_DATE + timedelta(days=13)),
    3: (BASE_DATE + timedelta(days=14), BASE_DATE + timedelta(days=20)),
    4: (BASE_DATE + timedelta(days=21), BASE_DATE + timedelta(days=27)),
}


def date_str(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ──────────────────────────────────────────────
# Leantime JSON-RPC Client
# ──────────────────────────────────────────────

class LeantimeClient:
    def __init__(self, base_url: str, api_key: str):
        self.base_url = base_url.rstrip("/")
        self.api_url = f"{self.base_url}/api/jsonrpc"
        self.session = requests.Session()
        self.session.headers["x-api-key"] = api_key
        self.session.headers["Content-Type"] = "application/json"
        self._rpc_id = 0

    def _call(self, method: str, params: dict = None) -> dict:
        self._rpc_id += 1
        payload = {
            "jsonrpc": "2.0",
            "method": method,
            "id": self._rpc_id,
        }
        if params:
            payload["params"] = params

        resp = self.session.post(self.api_url, json=payload)
        resp.raise_for_status()
        result = resp.json()

        if "error" in result:
            raise Exception(f"RPC Error [{method}]: {result['error']}")

        return result.get("result")

    def test_connection(self):
        try:
            result = self._call("leantime.rpc.projects.getAll")
            n = len(result) if result else 0
            print(f"  Conexao OK. Projetos existentes: {n}")
            return True
        except Exception as e:
            print(f"  Erro: {e}")
            return False

    def create_project(self, name: str, details: str = "",
                       start: str = None, end: str = None,
                       client_id: int = 0) -> int:
        params = {
            "name": name,
            "details": details,
            "clientId": client_id or "",
        }
        if start:
            params["start"] = start
        if end:
            params["end"] = end

        result = self._call("leantime.rpc.projects.addProject", params)
        pid = int(result) if result else 0
        print(f"  Projeto: {name} (id={pid})")
        return pid

    def create_milestone(self, project_id: int, headline: str,
                         edit_from: str = None, edit_to: str = None,
                         color: str = "#1b75bb") -> int:
        params = {
            "headline": headline[:255],
            "projectId": project_id,
            "tags": color,
        }
        if edit_from:
            params["editFrom"] = edit_from
        if edit_to:
            params["editTo"] = edit_to

        result = self._call("leantime.rpc.tickets.quickAddMilestone", params)
        mid = int(result) if result else 0
        return mid

    def create_ticket(self, project_id: int, headline: str,
                      description: str = "", priority: int = 3,
                      status: int = 3, ticket_type: str = "task",
                      tags: str = "", milestone_id: int = None,
                      depending_ticket_id: int = None,
                      date_to_finish: str = None,
                      edit_from: str = None, edit_to: str = None,
                      plan_hours: float = None,
                      acceptance_criteria: str = None,
                      story_points: int = None) -> int:
        params = {
            "headline": headline[:255],
            "description": description,
            "projectId": project_id,
            "priority": str(priority),
            "status": str(status),
            "type": ticket_type,
            "tags": tags,
        }
        if milestone_id:
            params["milestoneid"] = str(milestone_id)
        if depending_ticket_id:
            params["dependingTicketId"] = str(depending_ticket_id)
        if date_to_finish:
            params["dateToFinish"] = date_to_finish
        if edit_from:
            params["editFrom"] = edit_from
        if edit_to:
            params["editTo"] = edit_to
        if plan_hours:
            params["planHours"] = str(plan_hours)
        if acceptance_criteria:
            params["acceptanceCriteria"] = acceptance_criteria
        if story_points:
            params["storypoints"] = str(story_points)

        result = self._call("leantime.rpc.tickets.addTicket", params)
        return int(result) if result else 0

    def patch_ticket(self, ticket_id: int, fields: dict):
        params = {"id": ticket_id}
        params.update(fields)
        self._call("leantime.rpc.tickets.patchTicket", params)


# ──────────────────────────────────────────────
# HTML Description Builder
# ──────────────────────────────────────────────

def build_html(sections: list) -> str:
    """Constroi HTML rico a partir de lista de (tipo, titulo, conteudo)."""
    html = ""
    for sec_type, title, content in sections:
        if not content:
            continue
        if sec_type == "meta":
            html += f'<table style="width:100%;border-collapse:collapse;margin-bottom:12px;">\n'
            for label, value in content:
                if value:
                    html += f'<tr><td style="padding:4px 8px;font-weight:bold;width:140px;background:#f5f5f5;border:1px solid #ddd;">{label}</td>'
                    html += f'<td style="padding:4px 8px;border:1px solid #ddd;">{value}</td></tr>\n'
            html += '</table>\n'
        elif sec_type == "section":
            html += f'<h3 style="margin-top:16px;color:#333;border-bottom:1px solid #ddd;padding-bottom:4px;">{title}</h3>\n'
            # Converter newlines em <br> e preservar listas
            text = str(content).replace("\n", "<br>\n")
            html += f'<p>{text}</p>\n'
        elif sec_type == "checklist":
            html += f'<h3 style="margin-top:16px;color:#333;">{title}</h3>\n<ul>\n'
            for item in content:
                html += f'<li>{item}</li>\n'
            html += '</ul>\n'
    return html


# ──────────────────────────────────────────────
# Excel Parsers — versao rica
# ──────────────────────────────────────────────

def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def parse_prazo_to_date(prazo: str) -> tuple:
    """Converte prazo textual em (edit_from, edit_to, date_to_finish)."""
    p = prazo.upper().strip()
    if not p or p == "—":
        return None, None, None

    today = BASE_DATE
    if p == "HOJE":
        return date_str(today), date_str(today), date_str(today)
    elif "48H" in p:
        end = today + timedelta(days=2)
        return date_str(today), date_str(end), date_str(end)
    elif "72H" in p:
        end = today + timedelta(days=3)
        return date_str(today), date_str(end), date_str(end)
    elif p.startswith("SEM"):
        try:
            week = int(p.replace("SEM", "").strip())
            start, end = WEEK_DATES.get(week, (today, today + timedelta(days=6)))
            return date_str(start), date_str(end), date_str(end)
        except ValueError:
            return None, None, None
    elif "ASSEMBLEIA" in p or "PÓS" in p.upper():
        return date_str(ASSEMBLEIA_DATE), date_str(ASSEMBLEIA_DATE + timedelta(days=7)), date_str(ASSEMBLEIA_DATE + timedelta(days=7))
    elif "DIÁRIO" in p or "DIARIO" in p:
        return date_str(today), date_str(today + timedelta(days=30)), None
    return None, None, None


def parse_tarefas_por_pessoa(wb) -> list:
    ws = wb["16_TAREFAS_POR_PESSOA"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < 11:
            continue
        cell_id = safe_str(row_list[1])
        if not (cell_id and "-" in cell_id and len(cell_id) <= 5):
            continue

        responsavel = safe_str(row_list[2])
        tarefa = safe_str(row_list[3])
        prazo = safe_str(row_list[4])
        status_text = safe_str(row_list[5])
        prioridade = safe_str(row_list[6])
        entregavel = safe_str(row_list[7])
        depende_de = safe_str(row_list[8])
        obs = safe_str(row_list[9])

        titulo_lines = tarefa.split("\n")
        titulo = f"[{cell_id}] {titulo_lines[0]}"
        detalhes = "\n".join(titulo_lines[1:]) if len(titulo_lines) > 1 else ""

        # Build HTML
        html_sections = [
            ("meta", "", [
                ("ID", cell_id),
                ("Responsavel", responsavel),
                ("Prazo", prazo),
                ("Status", status_text),
                ("Prioridade", prioridade),
                ("Depende de", depende_de if depende_de != "—" else ""),
            ]),
        ]
        if detalhes:
            html_sections.append(("section", "O que fazer", detalhes))
        if entregavel:
            # Parse entregaveis as checklist
            items = [e.strip() for e in entregavel.split("\n") if e.strip()]
            html_sections.append(("checklist", "Entregaveis", items))
        if obs:
            html_sections.append(("section", "Observacoes", obs))

        description = build_html(html_sections)

        # Acceptance criteria from entregaveis
        acceptance = entregavel if entregavel else ""

        edit_from, edit_to, date_finish = parse_prazo_to_date(prazo)
        status_code = STATUS_MAP.get(status_text, 3)
        prio_code = PRIORITY_MAP.get(prioridade, 3)

        # Determine task type
        task_type = "task"
        if cell_id.startswith("J-"):
            task_type = "story"  # Jair's tasks are strategic decisions

        tags_list = [responsavel]
        if prioridade and "URGENTE" in prioridade.upper():
            tags_list.append("URGENTE")
        if "Bloqueado" in status_text or "⚫" in status_text:
            tags_list.append("BLOQUEADO")

        tasks.append({
            "id_code": cell_id,
            "title": titulo,
            "description": description,
            "status": status_code,
            "priority": prio_code,
            "type": task_type,
            "tags": ",".join(filter(None, tags_list)),
            "depends_on_code": depende_de if depende_de != "—" else "",
            "edit_from": edit_from,
            "edit_to": edit_to,
            "date_to_finish": date_finish,
            "acceptance_criteria": acceptance,
        })

    return tasks


def parse_cronograma(wb) -> list:
    ws = wb["02_CRONOGRAMA_GERAL"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < 10:
            continue
        num = safe_str(row_list[1])
        if not (num and num.isdigit()):
            continue

        atividade = safe_str(row_list[2])
        responsavel = safe_str(row_list[3])
        prerequisito = safe_str(row_list[4])
        obs = safe_str(row_list[9])

        semanas_ativas = []
        for i, sem in enumerate([5, 6, 7, 8], 1):
            if row_list[sem]:
                semanas_ativas.append(i)

        # Calcular datas do Gantt
        if semanas_ativas:
            first_week = min(semanas_ativas)
            last_week = max(semanas_ativas)
            edit_from = date_str(WEEK_DATES[first_week][0])
            edit_to = date_str(WEEK_DATES[last_week][1])
        else:
            edit_from, edit_to = None, None

        html_sections = [
            ("meta", "", [
                ("Atividade #", num),
                ("Responsavel", responsavel),
                ("Pre-requisito", prerequisito if prerequisito != "—" else ""),
                ("Periodo", ", ".join(f"Semana {s}" for s in semanas_ativas)),
            ]),
        ]
        if obs:
            html_sections.append(("section", "Observacoes", obs))

        # Depends on previous item
        depends_code = ""
        if prerequisito and prerequisito != "—":
            depends_code = prerequisito  # "Item 1", "Itens 1-4", etc.

        tasks.append({
            "id_code": f"CRON-{num}",
            "title": f"{num}. {atividade}",
            "description": build_html(html_sections),
            "status": 3,
            "priority": 3,
            "type": "task",
            "tags": responsavel,
            "depends_on_code": depends_code,
            "edit_from": edit_from,
            "edit_to": edit_to,
            "date_to_finish": edit_to,
        })

    return tasks


def parse_checklist_cartorio(wb) -> list:
    ws = wb["03_CHECKLIST_CARTORIO"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < 7:
            continue
        num = row_list[0]
        if not (num and isinstance(num, (int, float))):
            continue

        num = int(num)
        categoria = safe_str(row_list[1])
        item = safe_str(row_list[2])
        descricao = safe_str(row_list[3])
        status_text = safe_str(row_list[4])
        responsavel = safe_str(row_list[5])
        obs = safe_str(row_list[6])

        done = status_text.lower() in ("concluído", "concluido")
        status_code = 0 if done else 3

        html_sections = [
            ("meta", "", [
                ("Item #", str(num)),
                ("Categoria", categoria),
                ("Status", status_text),
                ("Responsavel", responsavel if responsavel != "—" else ""),
            ]),
        ]
        if descricao:
            html_sections.append(("section", "Exigencia", descricao))
        if obs:
            html_sections.append(("section", "Observacoes", obs))

        tags_parts = [categoria.split(")")[0].replace("(", "").strip()] if ")" in categoria else [categoria]
        if responsavel and responsavel != "—":
            tags_parts.append(responsavel)

        tasks.append({
            "id_code": f"CART-{num}",
            "title": f"#{num} {item}",
            "description": build_html(html_sections),
            "status": status_code,
            "priority": 3,
            "type": "task",
            "tags": ",".join(filter(None, tags_parts)),
        })

    return tasks


def parse_pendencias(wb) -> list:
    ws = wb["14_PENDENCIAS_INSUMO"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < 7:
            continue
        num_raw = safe_str(row_list[1])
        if not num_raw:
            continue
        try:
            num = int(float(num_raw))
        except (ValueError, TypeError):
            continue
        if num < 1 or num > 99:
            continue

        pendencia = safe_str(row_list[2])
        fonte = safe_str(row_list[3])
        criticidade = safe_str(row_list[4])
        responsavel = safe_str(row_list[5])
        obs = safe_str(row_list[6])
        if not pendencia:
            continue

        done = "CONCLU" in criticidade.upper() or "CONCLU" in obs.upper()[:20]

        html_sections = [
            ("meta", "", [
                ("Pendencia #", str(num)),
                ("Fonte esperada", fonte),
                ("Criticidade", criticidade),
                ("Responsavel", responsavel),
            ]),
        ]
        if obs:
            html_sections.append(("section", "Observacoes", obs))

        prio = 2 if "ALTA" in criticidade.upper() else 3
        is_blocker = "BLOQ" in criticidade.upper() or "BLOQ" in obs.upper()
        tags_parts = [responsavel]
        if is_blocker:
            tags_parts.append("BLOQUEANTE")

        tasks.append({
            "id_code": f"PEND-{num}",
            "title": f"Pendencia #{num}: {pendencia}",
            "description": build_html(html_sections),
            "status": 0 if done else (1 if is_blocker else 3),
            "priority": prio,
            "type": "bug",
            "tags": ",".join(filter(None, tags_parts)),
        })

    return tasks


def parse_riscos(wb) -> list:
    ws = wb["12_RISCOS_CONTROLES"]
    tasks = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < 8:
            continue
        num_raw = safe_str(row_list[1])
        if not num_raw:
            continue
        try:
            num = int(float(num_raw))
        except (ValueError, TypeError):
            continue
        if num < 1 or num > 99:
            continue

        risco = safe_str(row_list[2])
        if not risco or risco == "Risco":
            continue

        probabilidade = safe_str(row_list[3])
        impacto = safe_str(row_list[4])
        controle = safe_str(row_list[5])
        responsavel = safe_str(row_list[6])

        html_sections = [
            ("meta", "", [
                ("Risco #", str(num)),
                ("Probabilidade", probabilidade),
                ("Impacto", impacto),
                ("Responsavel", responsavel),
            ]),
        ]
        if controle:
            html_sections.append(("section", "Controle / Mitigacao", controle))

        prio = 2 if impacto in ("Alto", "Crítico", "Critico") else 3

        tasks.append({
            "id_code": f"RISCO-{num}",
            "title": f"Risco #{num}: {risco}",
            "description": build_html(html_sections),
            "status": 3,
            "priority": prio,
            "type": "bug",
            "tags": ",".join(filter(None, [responsavel, f"Impacto-{impacto}"])),
        })

    return tasks


def parse_cnpj_redesim(wb) -> list:
    ws = wb["04_CNPJ_REDESIM_RFB"]
    tasks = []
    current_phase = ""
    prev_code = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < 7:
            continue
        passo = safe_str(row_list[1])

        if passo.startswith("FASE"):
            current_phase = passo
            prev_code = None
            continue

        if not (passo and "." in passo):
            continue
        try:
            float(passo)
        except ValueError:
            continue

        oque_fazer = safe_str(row_list[2])
        quem = safe_str(row_list[3])
        onde = safe_str(row_list[4])
        obs = safe_str(row_list[6])

        code = f"CNPJ-{passo}"

        html_sections = [
            ("meta", "", [
                ("Passo", passo),
                ("Fase", current_phase),
                ("Quem", quem),
                ("Onde / Sistema", onde),
            ]),
        ]
        if oque_fazer:
            html_sections.append(("section", "O que fazer", oque_fazer))
        if obs:
            html_sections.append(("section", "Observacoes", obs))

        tasks.append({
            "id_code": code,
            "title": f"[{passo}] {oque_fazer[:200]}",
            "description": build_html(html_sections),
            "status": 3,
            "priority": 3,
            "type": "task",
            "tags": quem,
            "depends_on_code": prev_code or "",  # Sequential dependency
        })
        prev_code = code

    return tasks


# ──────────────────────────────────────────────
# Dependency Resolver
# ──────────────────────────────────────────────

def resolve_dependencies(all_tasks: dict, ticket_ids: dict):
    """
    Resolve dependencias textuais para IDs reais de tickets.
    Retorna lista de (ticket_id, depending_ticket_id).
    """
    deps = []

    for section_name, tasks in all_tasks.items():
        for t in tasks:
            dep_code = t.get("depends_on_code", "")
            if not dep_code:
                continue

            task_ticket_id = ticket_ids.get(t["id_code"])
            if not task_ticket_id:
                continue

            # Direct code match: "J-01", "C-01", "CNPJ-1.1"
            if dep_code in ticket_ids:
                deps.append((task_ticket_id, ticket_ids[dep_code]))
                continue

            # Extract task codes from text like "Assembleia realizada + Diretoria eleita (J-01)"
            import re
            codes_found = re.findall(r'[A-Z]+-\d+(?:\.\d+)?', dep_code)
            if codes_found:
                # Use the first found code as predecessor
                for code in codes_found:
                    if code in ticket_ids:
                        deps.append((task_ticket_id, ticket_ids[code]))
                        break
                continue

            # Cronograma: "Item 1", "Itens 1-4", "Itens 5-8"
            item_match = re.search(r'[Ii]te[mn]s?\s+(\d+)(?:-(\d+))?', dep_code)
            if item_match:
                last_item = item_match.group(2) or item_match.group(1)
                cron_code = f"CRON-{last_item}"
                if cron_code in ticket_ids:
                    deps.append((task_ticket_id, ticket_ids[cron_code]))
                continue

    return deps


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

# Milestone definitions for the main project
MILESTONES = [
    {
        "name": "Fase 1 — Preparacao",
        "color": "#3498db",
        "edit_from": date_str(WEEK_DATES[1][0]),
        "edit_to": date_str(WEEK_DATES[2][1]),
    },
    {
        "name": "Fase 2 — Assembleia",
        "color": "#27ae60",
        "edit_from": date_str(WEEK_DATES[3][0]),
        "edit_to": date_str(ASSEMBLEIA_DATE),
    },
    {
        "name": "Fase 3 — Registro e CNPJ",
        "color": "#e67e22",
        "edit_from": date_str(ASSEMBLEIA_DATE + timedelta(days=1)),
        "edit_to": date_str(ASSEMBLEIA_DATE + timedelta(days=21)),
    },
]

PROJECT_DEFS = {
    "Tarefas Operacionais": {
        "details": "<p>Tarefas do dia a dia por responsavel — Caina (estagiario), Jair (coordenador), Advogado, Brava Capital. Extraido da aba 16_TAREFAS_POR_PESSOA do Manual Operacional v2.1.</p>",
        "start": date_str(BASE_DATE),
        "end": date_str(ASSEMBLEIA_DATE + timedelta(days=30)),
    },
    "Cronograma Geral": {
        "details": "<p>15 atividades sequenciais em 4 semanas — da coleta de dados ate abertura de conta bancaria. Extraido da aba 02_CRONOGRAMA_GERAL. Use a vista Gantt para ver a timeline.</p>",
        "start": date_str(WEEK_DATES[1][0]),
        "end": date_str(WEEK_DATES[4][1]),
    },
    "Checklist Cartorio": {
        "details": "<p>51 itens de conformidade documental para registro da associacao no Cartorio de RCPJ de Itapema/SC. Fonte: registrositapema.com.br + Codigo de Normas CGJ-SC. Aba 03_CHECKLIST_CARTORIO.</p>",
        "start": date_str(BASE_DATE),
        "end": date_str(ASSEMBLEIA_DATE + timedelta(days=14)),
    },
    "Processo CNPJ-REDESIM": {
        "details": "<p>33 passos sequenciais para abertura do CNPJ via REDESIM — do pre-requisito ao CNPJ ativo. 5 fases: Pre-requisitos, Viabilidade, DBE, Pos-registro, Manutencao. Aba 04_CNPJ_REDESIM_RFB.</p>",
        "start": date_str(ASSEMBLEIA_DATE),
        "end": date_str(ASSEMBLEIA_DATE + timedelta(days=60)),
    },
    "Pendencias e Bloqueios": {
        "details": "<p>10 itens pendentes de insumo que bloqueiam progresso. Resolver ANTES de avancar. Aba 14_PENDENCIAS_INSUMO.</p>",
        "start": date_str(BASE_DATE),
        "end": date_str(ASSEMBLEIA_DATE),
    },
    "Riscos e Controles": {
        "details": "<p>Matriz de 10 riscos identificados com probabilidade, impacto e controles de mitigacao. Aba 12_RISCOS_CONTROLES.</p>",
        "start": date_str(BASE_DATE),
        "end": date_str(ASSEMBLEIA_DATE + timedelta(days=60)),
    },
}


def main():
    parser = argparse.ArgumentParser(description="Importar tarefas do Excel para Leantime")
    parser.add_argument("--url", required=True, help="URL do Leantime")
    parser.add_argument("--api-key", required=True, help="API key")
    parser.add_argument("--excel", default=str(EXCEL_FILE), help="Caminho do Excel")
    parser.add_argument("--dry-run", action="store_true", help="Apenas mostrar")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  IMPORTADOR LEANTIME — Manual Operacional Associacao")
    print(f"{'='*60}")

    # 1. Carregar Excel
    print(f"\n1. Carregando Excel: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, data_only=True)

    # 2. Parsear
    print("\n2. Extraindo tarefas...")
    sections = {
        "Tarefas Operacionais": parse_tarefas_por_pessoa(wb),
        "Cronograma Geral": parse_cronograma(wb),
        "Checklist Cartorio": parse_checklist_cartorio(wb),
        "Processo CNPJ-REDESIM": parse_cnpj_redesim(wb),
        "Pendencias e Bloqueios": parse_pendencias(wb),
        "Riscos e Controles": parse_riscos(wb),
    }

    total = sum(len(v) for v in sections.values())
    for name, tasks in sections.items():
        done = sum(1 for t in tasks if t.get("status") == 0)
        deps = sum(1 for t in tasks if t.get("depends_on_code"))
        print(f"   {name}: {len(tasks)} tarefas ({done} concluidas, {deps} com dependencias)")
    print(f"   TOTAL: {total} tarefas")

    if args.dry_run:
        print("\n=== DRY RUN ===\n")
        prio_names = {1: "CRIT", 2: "HIGH", 3: "MED", 4: "LOW", 5: "LOWST"}
        status_names = {0: "DONE", 1: "BLOCK", 3: "NEW", 4: "INPROG", -1: "ARCH"}
        for name, tasks in sections.items():
            print(f"\n--- {name} ({len(tasks)} tarefas) ---")
            for t in tasks:
                s = status_names.get(t.get("status", 3), "?")
                p = prio_names.get(t.get("priority", 3), "?")
                dep = f" ← {t['depends_on_code']}" if t.get("depends_on_code") else ""
                dates = ""
                if t.get("edit_from"):
                    dates = f" [{t['edit_from'][:10]}..{t.get('edit_to', '')[:10]}]"
                print(f"  [{s}] P{t.get('priority',3)} {t['title'][:65]}{dep}{dates}")
        return

    # 3. Conectar
    print(f"\n3. Conectando ao Leantime: {args.url}")
    client = LeantimeClient(args.url, args.api_key)
    if not client.test_connection():
        print("FALHA. Verifique URL e API key.")
        sys.exit(1)

    # 4. Criar projetos e milestones
    print("\n4. Criando projetos e milestones...")
    project_ids = {}
    milestone_ids = {}

    for name in sections:
        pdef = PROJECT_DEFS.get(name, {})
        pid = client.create_project(
            name=name,
            details=pdef.get("details", ""),
            start=pdef.get("start"),
            end=pdef.get("end"),
        )
        project_ids[name] = pid

        # Create milestones for Tarefas Operacionais
        if name == "Tarefas Operacionais":
            for m in MILESTONES:
                mid = client.create_milestone(
                    pid, m["name"],
                    edit_from=m["edit_from"],
                    edit_to=m["edit_to"],
                    color=m["color"],
                )
                milestone_ids[m["name"]] = mid
                print(f"    Milestone: {m['name']} (id={mid})")

    # 5. Criar tarefas
    print("\n5. Importando tarefas...")
    ticket_ids = {}  # code -> ticket_id
    stats = {"created": 0, "done": 0}

    for name, tasks in sections.items():
        pid = project_ids.get(name)
        if not pid:
            continue

        for t in tasks:
            # Determine milestone for operational tasks
            ms_id = None
            code = t.get("id_code", "")
            if name == "Tarefas Operacionais":
                # C-01..C-05 and J-01..J-02 are Phase 1 (Preparation)
                if any(code.startswith(p) for p in ["C-01", "C-02", "C-03", "C-04", "C-05",
                                                      "J-01", "J-02", "J-03", "J-04",
                                                      "A-01", "A-02", "A-04",
                                                      "CT-01", "B-01", "B-02"]):
                    ms_id = milestone_ids.get("Fase 1 — Preparacao")
                elif any(code.startswith(p) for p in ["C-06", "J-05", "C-08"]):
                    ms_id = milestone_ids.get("Fase 2 — Assembleia")
                elif any(code.startswith(p) for p in ["CT-02", "A-03"]):
                    ms_id = milestone_ids.get("Fase 3 — Registro e CNPJ")

            tid = client.create_ticket(
                project_id=pid,
                headline=t["title"],
                description=t.get("description", ""),
                priority=t.get("priority", 3),
                status=t.get("status", 3),
                ticket_type=t.get("type", "task"),
                tags=t.get("tags", ""),
                milestone_id=ms_id,
                date_to_finish=t.get("date_to_finish"),
                edit_from=t.get("edit_from"),
                edit_to=t.get("edit_to"),
                acceptance_criteria=t.get("acceptance_criteria"),
            )

            if code:
                ticket_ids[code] = tid

            stats["created"] += 1
            if t.get("status") == 0:
                stats["done"] += 1

        done = sum(1 for t in tasks if t.get("status") == 0)
        print(f"  {name}: {len(tasks)} tarefas criadas ({done} concluidas)")

    # 6. Resolver dependencias
    print("\n6. Resolvendo dependencias...")
    deps = resolve_dependencies(sections, ticket_ids)
    dep_count = 0
    for tid, dep_tid in deps:
        try:
            client.patch_ticket(tid, {"dependingTicketId": str(dep_tid)})
            dep_count += 1
        except Exception as e:
            pass  # best effort

    print(f"  {dep_count} dependencias resolvidas de {len(deps)} encontradas")

    # 7. Resumo
    print(f"\n{'='*60}")
    print(f"  IMPORTACAO CONCLUIDA!")
    print(f"  Projetos: {len(project_ids)}")
    print(f"  Milestones: {len(milestone_ids)}")
    print(f"  Tarefas: {stats['created']} ({stats['done']} concluidas)")
    print(f"  Dependencias: {dep_count}")
    print(f"\n  Aceda em: {args.url}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
